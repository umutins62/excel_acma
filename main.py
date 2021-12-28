import openpyxl
from PyQt5.QtWidgets import QTableWidget, QApplication, QPushButton, QVBoxLayout,QLineEdit, QHeaderView, QTableWidgetItem, QLabel, QHBoxLayout
import sys
import pandas as pd
from PyQt5.QtWidgets import (QWidget,
                             QFileDialog)
from PyQt5 import QtGui
from openpyxl.styles import Side, Border, Alignment, numbers


class excelac(QWidget):
    def __init__(self):
        super().__init__()
        self.setUI()

    def setUI(self):

        self.setWindowTitle("YAPIM İŞİ HAKEDİŞ TAKİP")
        self.setGeometry(600, 300, 1110, 300)
        self.setWindowIcon(QtGui.QIcon('MUK.png'))

        self.ac=QPushButton("YM Aç")
        self.ac.clicked.connect(self.dosya_ac)



        self.kaydet=QPushButton("Hesapla")
        self.kaydet.clicked.connect(self.hesapla)

        self.kaydet1=QPushButton("Kaydet")
        self.kaydet1.clicked.connect(self.kaydetexc)

        self.mukayese=QPushButton("Mukayese Al")
        self.mukayese.clicked.connect(self.mukayeseal)



        self.hakedis=QLineEdit()
        self.hakedis.setPlaceholderText("Hakediş No")

        self.hakedis_Ekle=QPushButton("Hakediş Ekle")
        self.hakedis_Ekle.clicked.connect(self.hakedisEkle)


        self.toplam=QLabel("Toplam Tutar: ")
        self.toplam.setStyleSheet("background-color: black;border: 1px solid black;color : chartreuse;font-size: 10pt")
        self.kalan=QLabel("Kalan Tutar: ")
        self.kalan.setStyleSheet("background-color: black;border: 1px solid black;color : chartreuse;font-size: 10pt")
        self.islem_göstergesi=QLabel("Çalışıyor... ")
        self.islem_göstergesi.setStyleSheet("background-color: black;border: 1px solid black;color : chartreuse;font-size: 10pt")

        self.tableWidget = QTableWidget()
        self.tableWidget.setAlternatingRowColors(True)
        self.tableWidget.setColumnCount(1)

        self.tableWidget.horizontalHeader().setCascadingSectionResizes(False)
        self.tableWidget.horizontalHeader().setSortIndicatorShown(False)
        self.tableWidget.horizontalHeader().setStretchLastSection(True)
        self.tableWidget.horizontalHeader().setVisible(False)
        self.tableWidget.verticalHeader().setVisible(False)
        self.tableWidget.verticalHeader().setCascadingSectionResizes(False)
        self.tableWidget.verticalHeader().setStretchLastSection(False)
        # self.tableWidget.setHorizontalHeaderLabels(("Sıra No","Tarihi","Konusu", "Mahkeme"
        #                                             ,"Ücreti","Net\nKalan", "Teslim\nTarihi","Hazırlanış\nSüresi"))
        # self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        style = ":section {""background-color: silver ; }"
        self.tableWidget.horizontalHeader().setStyleSheet(style)
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # self.tableWidget.itemChanged.connect(self.hesapla)


        h2box=QVBoxLayout()
        v1box=QHBoxLayout()
        v2box=QHBoxLayout()


        v2box.addWidget(self.islem_göstergesi,80)
        # v2box.addStretch()
        v2box.addWidget(self.toplam,10)
        v2box.addWidget(self.kalan,10)

        v1box.addWidget(self.hakedis)
        v1box.addWidget(self.hakedis_Ekle)

        v1box.addWidget(self.kaydet1)
        v1box.addWidget(self.mukayese)
        v1box.addStretch(4)
        v1box.addWidget(self.kaydet)
        v1box.addWidget(self.ac)

        h2box.addLayout(v1box)
        h2box.addWidget(self.tableWidget)
        h2box.addLayout(v2box)

        self.setLayout(h2box)

        self.show()

    def mukayeseal(self):
        pass

    def kaydetexc(self):
        # try:
            print(self.df.dtypes)
            self.df.to_excel("mukayeseli.xlsx")
            filename='mukayeseli.xlsx'

            wb = openpyxl.load_workbook(filename,data_only=True)
            ws = wb['Sheet1']
            ws.delete_cols(idx=1)




            thin = Side(border_style="thin", color="000000")#Border style, color
            border = Border(left=thin, right=thin, top=thin, bottom=thin)#Position of border

            row_count = ws.max_row
            column_count = ws.max_column
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))

                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.1
                ws.column_dimensions[column].width = adjusted_width

            ws["G"+str(row_count+1)]=str(round(self.Total,2))



            for col in ws.columns:
                for cell in col:
                    if len(str(cell.value)) > 0:
                        cell.alignment = Alignment(horizontal='center')





            for row in ws["A1:I"+str(row_count+1)]:
                for cell in row:
                    cell.border = border#A5:D6 range cell setting border



            ws.merge_cells(start_row=row_count+1, start_column=1, end_row=row_count+1, end_column=6)
            ws["A"+str(row_count+1)]="TOPLAM"
            ws["A"+str(row_count+1)].alignment = Alignment(horizontal='right')
            ws["A"+str(row_count+1)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE


            ws["I"+str(row_count+1)]=str(round(self.Total1))
            ws["I"+str(row_count+1)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE





            wb.save(filename)
            self.islem_göstergesi.setText("Excel'e aktarım başarılı")

        # except:
        #     self.islem_göstergesi.setText("Excel'e aktarım yapılamadı")


    def hesapla(self):

        try:

            rows=self.tableWidget.rowCount()

            list2=[]

            for col in self.df.columns:
                list2.append(col)
            Mi=int(list2.index(list2[-1]))
            BFi=int(list2.index(list2[-2]))

            for i in range(rows):
                BF = float(self.tableWidget.item(i, 5).text().replace(",","."))
                M = float(self.tableWidget.item(i, BFi).text().replace(",","."))

                S=round(BF*M,2)
                self.tableWidget.setItem(i, Mi, QTableWidgetItem(str(S)))
                self.df.at[i,list2[-1]]=S
                self.df.at[i,list2[-2]]=M




            self.Total1 = self.df[list2[-1]].sum()
            self.kalan.setText("Kalan Tutar: "+format(round(self.Total1,2), ",")+" TL")


            self.tableWidget.setColumnCount(len(list2))
            self.tableWidget.setHorizontalHeaderLabels(list2)
            self.tableWidget.setColumnCount(len(self.df.columns))
            self.tableWidget.setRowCount(len(self.df.index))

            for i in range(len(self.df.index)):
                for j in range(len(self.df.columns)):
                    self.tableWidget.setItem(i, j, QTableWidgetItem(str(self.df.iat[i, j])))
                    self.tableWidget.horizontalHeader().setVisible(True)

            self.tableWidget.resizeColumnsToContents()
            self.tableWidget.resizeRowsToContents()

            self.islem_göstergesi.setText("Hesap İşlemi  başarılı")




        except:
            self.islem_göstergesi.setText("Hesap İşlemi  yapılamadı!")

    def hakedisEkle(self):
        list=[]
        list1=[]

        for col in self.df.columns:
            list.append(col)
        hakedisadi=self.hakedis.text()

        self.df.insert(len(list), "HAKEDİŞ MİKTARI "+hakedisadi,0.0,  True)
        self.df.insert(len(list)+1, "HAKEDİŞ TUTARI "+hakedisadi,0.0,  True)
        self.df.rename(columns={"8": "HAKEDİŞ MİKTARI "+str(hakedisadi), "9":"HAKEDİŞ TUTARI "+str(hakedisadi)})


        for col1 in self.df.columns:
            list1.append(col1)

        self.tableWidget.setColumnCount(len(list1))

        self.tableWidget.setHorizontalHeaderLabels(list1)


        self.tableWidget.setColumnCount(len(self.df.columns))
        self.tableWidget.setRowCount(len(self.df.index))

        for i in range(len(self.df.index)):
            for j in range(len(self.df.columns)):
                self.tableWidget.setItem(i, j, QTableWidgetItem(str(self.df.iat[i, j])))
                self.tableWidget.horizontalHeader().setVisible(True)

        self.tableWidget.resizeColumnsToContents()
        self.tableWidget.resizeRowsToContents()


    def dosya_ac(self):
        try:
            fileName = QFileDialog.getOpenFileName(self, "Dosya Aç",
                                                   "/Excel Şeç",
                                                   "Excel (*.xls *.xlsx *.xlsm)")


            self.df = pd.read_excel (str(fileName[0]))

            list=[]

            for col in self.df.columns:
                list.append(col)

            print(self.df)


            self.Total = self.df[list[6]].sum()
            self.toplam.setText("Toplam Tutar: "+format(round(self.Total,2), ",")+" TL")

            self.tableWidget.setColumnCount(len(list))

            self.tableWidget.setHorizontalHeaderLabels(list)

            self.tableWidget.setColumnCount(len(self.df.columns))
            self.tableWidget.setRowCount(len(self.df.index))

            for i in range(len(self.df.index)):
                for j in range(len(self.df.columns)):
                    self.tableWidget.setItem(i, j, QTableWidgetItem(str(self.df.iat[i, j])))
                    self.tableWidget.horizontalHeader().setVisible(True)

            self.tableWidget.resizeColumnsToContents()
            self.tableWidget.resizeRowsToContents()


        except:
            self.islem_göstergesi.setText("Excelden veri çekerken bir hata oluştu!")



if __name__ == "__main__":
    app = QApplication(sys.argv)
    pencere = excelac()
    sys.exit(app.exec())
